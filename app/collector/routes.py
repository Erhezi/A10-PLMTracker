import re
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload
from datetime import date, datetime, timedelta
from openpyxl import load_workbook

from .. import db
from ..models import now_ny_naive
from ..models.relations import (
    ItemLink,
    ItemLinkWrike,
    ItemLinkArchived,
    ItemLinkDeleted,
    ItemGroup,
    PendingItems,
    ItemGroupConflictError,
    ConflictError,
)
from ..models.inventory import Item, ContractItem
from ..models.log import BurnRateRefreshJob
from ..utility.item_group import BatchValidationError
from ..utility.add_pairs import AddItemPairs
from ..utility.stage_transition import StageTransitionHelper
from .batch_service import (
    apply_stage,
    apply_wrike,
    apply_go_live,
    summarize_results,
)

"""Stage name definitions (canonical only)."""

ALLOWED_STAGES = StageTransitionHelper.STAGES  # New canonical stage names (order meaningful for UI select)

def normalize_stage(value: str | None) -> str | None:  # kept for minimal refactor
    return value

SENTINEL_REPLACEMENTS = {"NO REPLACEMENT"}  # Dynamic PENDING*** codes handled separately
ITEM_CODE_PATTERN = re.compile(r"^[1-9]\d{5}$")
ALLOWED_UPLOAD_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")

bp = Blueprint("collector", __name__, url_prefix="")

@bp.route("/collect")
@login_required
def collect():
    # Fetch a small sample set to verify Item view connectivity
    sample_items = Item.query.order_by(Item.item).limit(25).all()
    from calendar import monthrange

    def add_months(d: date, months: int) -> date:
        m = d.month - 1 + months
        y = d.year + m // 12
        m = m % 12 + 1
        day = min(d.day, monthrange(y, m)[1])
        from datetime import date as _d
        return _d(y, m, day)

    today = date.today()
    min_date = add_months(today, -3)
    max_date = add_months(today, 6)
    
    # Import config for batch limits
    from flask import current_app
    max_per_side = current_app.config.get('MAX_BATCH_PER_SIDE', 6)
    max_combinations = max_per_side * max_per_side  # Calculate total combinations for display
    
    return render_template(
        "collector/collect.html",
        allowed_stages=ALLOWED_STAGES,
        sample_items=sample_items,
        date_min=min_date.isoformat(),
        date_max=max_date.isoformat(),
        max_combinations=max_combinations,
        max_per_side=max_per_side,
    )

@bp.route("/groups")
@login_required
def groups():
    # Avoid eager loading any relationships (they are not needed for groups table)
    items = (
        ItemLink.query
        .options(selectinload(ItemLink.wrike))
        .order_by(ItemLink.item_group.desc(), ItemLink.item)
        .all()
    )
    # Count rows currently flagged as deleted
    count_deleted = ItemLink.query.filter(ItemLink.stage == 'Deleted').count()
    count_completed = ItemLink.query.filter(ItemLink.stage == 'Tracking Completed').count()
    stage_transitions = {
        stage: sorted(StageTransitionHelper.allowed_targets(stage))
        for stage in ALLOWED_STAGES
    }
    stage_transitions["__new__"] = list(ALLOWED_STAGES)
    return render_template(
        "collector/groups.html",
        items=items,
        allowed_stages=ALLOWED_STAGES,
        count_deleted=count_deleted,
        count_completed=count_completed,
        stage_transitions=stage_transitions,
    )

@bp.post('/groups/clear-deleted')
@login_required
def clear_deleted():
    # Move rows marked Deleted into the dedicated history table before removal
    deleted_query = ItemLink.query.filter(ItemLink.stage == 'Deleted')
    records = deleted_query.all()
    if not records:
        flash('No rows with stage Deleted to remove', 'info')
        return redirect(url_for('collector.groups'))

    deleted_rows: list[ItemLinkDeleted] = []
    deleted_time = now_ny_naive()
    for record in records:
        deleted_rows.append(
            ItemLinkDeleted(
                item_group=record.item_group,
                item=record.item,
                replace_item=record.replace_item,
                mfg_part_num=record.mfg_part_num,
                manufacturer=record.manufacturer,
                item_description=record.item_description,
                repl_mfg_part_num=record.repl_mfg_part_num,
                repl_manufacturer=record.repl_manufacturer,
                repl_item_description=record.repl_item_description,
                stage=record.stage,
                expected_go_live_date=record.expected_go_live_date,
                create_dt=record.create_dt,
                update_dt=record.update_dt,
                item_link_id=record.pkid,
                deleted_dt=deleted_time,
            )
        )

    try:
        if deleted_rows:
            db.session.add_all(deleted_rows)
        deleted_query.delete(synchronize_session=False)
        db.session.commit()
    except Exception as exc:  # pragma: no cover - rollback safety
        db.session.rollback()
        flash(f'Clearing deleted rows failed: {exc}', 'danger')
        return redirect(url_for('collector.groups'))

    flash(f'Cleared {len(deleted_rows)} deleted item link(s)', 'success')
    return redirect(url_for('collector.groups'))

@bp.post('/groups/archive-completed')
@login_required
def archive_completed():
    # Move Tracking Completed rows into the archive table, preserving history
    completed_query = ItemLink.query.filter(ItemLink.stage == 'Tracking Completed')
    records = completed_query.all()
    if not records:
        flash('No completed item link rows to archive', 'info')
        return redirect(url_for('collector.groups'))

    archived_rows: list[ItemLinkArchived] = []
    archive_time = now_ny_naive()
    for record in records:
        archived_rows.append(
            ItemLinkArchived(
                item_group=record.item_group,
                item=record.item,
                replace_item=record.replace_item,
                mfg_part_num=record.mfg_part_num,
                manufacturer=record.manufacturer,
                item_description=record.item_description,
                repl_mfg_part_num=record.repl_mfg_part_num,
                repl_manufacturer=record.repl_manufacturer,
                repl_item_description=record.repl_item_description,
                stage=record.stage,
                expected_go_live_date=record.expected_go_live_date,
                create_dt=record.create_dt,
                update_dt=record.update_dt,
                item_link_id=record.pkid,
                archived_dt=archive_time,
            )
        )

    try:
        if archived_rows:
            db.session.add_all(archived_rows)
        completed_query.delete(synchronize_session=False)
        db.session.commit()
    except Exception as exc:  # pragma: no cover - safety rollback
        db.session.rollback()
        flash(f'Archiving failed: {exc}', 'danger')
        return redirect(url_for('collector.groups'))

    flash(f'Archived {len(archived_rows)} completed item link(s)', 'success')
    return redirect(url_for('collector.groups'))


@bp.route("/groups/<item>/<replace_item>/update", methods=["POST"])
@login_required
def update_item(item: str, replace_item: str):
    # The template may render Python None as the string 'None' (or similar) into
    # the data-replace-item attribute. Normalize common textual null
    # representations to actual None so SQL queries match NULL values.
    if isinstance(replace_item, str) and replace_item.lower() in ('none', 'null', 'nan', ''):
        replace_item = None
    record = (
        ItemLink.query
        .options(selectinload(ItemLink.wrike))
        .filter_by(item=item, replace_item=replace_item)
        .first_or_404()
    )
    stage = request.form.get("stage")
    expected_go_live_date = request.form.get("expected_go_live_date") or None
    wants_json = (request.args.get('ajax') == '1') or ('application/json' in request.headers.get('Accept','')) or request.headers.get('X-Requested-With') == 'fetch'
    wrike_inputs = {
        "wrike_id1": request.form.get("wrike_id1"),
        "wrike_id2": request.form.get("wrike_id2"),
        "wrike_id3": request.form.get("wrike_id3"),
    }

    normalized_wrike: dict[str, str | None] = {}
    for field, raw in wrike_inputs.items():
        value = (raw or "").strip()
        if value:
            if not (value.isdigit() and len(value) == 10):
                if wants_json:
                    return jsonify({"status": "error", "field": field, "message": "Wrike ID must be exactly 10 digits"}), 400
                flash("Wrike ID values must be 10 digits or left blank", "warning")
                return redirect(url_for("collector.groups"))
            normalized_wrike[field] = value
        else:
            normalized_wrike[field] = None

    decision = StageTransitionHelper.evaluate_transition(
        record.stage,
        stage,
        replace_item=record.replace_item,
    )

    if not decision.allowed:
        message = decision.reason or "Stage transition not permitted"
        if wants_json:
            return jsonify({"status": "error", "field": "stage", "message": message}), 400
        flash(message, "warning")
        return redirect(url_for("collector.groups"))

    record.stage = decision.final_stage
    # Parse date (YYYY-MM-DD) if provided
    if expected_go_live_date:
        try:
            from datetime import datetime
            record.expected_go_live_date = datetime.strptime(expected_go_live_date, "%Y-%m-%d").date()
        except ValueError:
            if wants_json:
                return jsonify({"status":"error","field":"expected_go_live_date","message":"Invalid date format; use YYYY-MM-DD"}), 400
            flash("Invalid date format; use YYYY-MM-DD", "warning")
    else:
        record.expected_go_live_date = None

    wrike_record = ItemLinkWrike.ensure_for_link(record)
    wrike_record.wrike_id1 = normalized_wrike["wrike_id1"]
    wrike_record.wrike_id2 = normalized_wrike["wrike_id2"]
    wrike_record.wrike_id3 = normalized_wrike["wrike_id3"]
    wrike_record.sync_from_item_link(record)

    record.update_dt = now_ny_naive()
    db.session.commit()
    if wants_json:
        # Also return current deleted count for client-side UI updates
        count_deleted = ItemLink.query.filter(ItemLink.stage == 'Deleted').count()
        count_completed = ItemLink.query.filter(ItemLink.stage == 'Tracking Completed').count()
        return jsonify({
            "status":"ok",
            "message":"ItemLink updated",
            "record":{
                "item": record.item,
                "replace_item": record.replace_item,
                "stage": record.stage,
                "expected_go_live_date": record.expected_go_live_date.isoformat() if record.expected_go_live_date else None,
                "wrike_id1": wrike_record.wrike_id1,
                "wrike_id2": wrike_record.wrike_id2,
                "wrike_id3": wrike_record.wrike_id3,
                "update_dt": record.update_dt.isoformat() if record.update_dt else None,
            },
            "count_deleted": count_deleted,
            "count_completed": count_completed,
            "transition_note": decision.reason,
        })
    flash("ItemLink updated", "success")
    if decision.reason:
        flash(decision.reason, "info")
    return redirect(url_for("collector.groups"))


@bp.post("/groups/batch/stage")
@login_required
def batch_update_stage():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    requested_stage = payload.get("stage")
    if not requested_stage:
        return jsonify({"status": "error", "message": "Missing stage"}), 400
    try:
        results = apply_stage(rows, requested_stage)
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    summary = summarize_results(results)
    return jsonify(summary)


@bp.post("/groups/batch/wrike/<field>")
@login_required
def batch_update_wrike(field: str):
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    value = payload.get("value")
    try:
        results = apply_wrike(rows, field, value)
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    summary = summarize_results(results)
    return jsonify(summary)


@bp.post("/groups/batch/go-live")
@login_required
def batch_update_go_live():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    date_value = payload.get("expected_go_live_date")
    try:
        results = apply_go_live(rows, date_value)
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    summary = summarize_results(results)
    return jsonify(summary)

@bp.route("/conflicts")
@login_required
def conflicts():
    limit = request.args.get("limit", type=int) or 200
    limit = max(1, min(limit, 1000))
    selected_type = request.args.get("type") or None

    type_counts_rows = (
        db.session.query(ConflictError.error_type, func.count(ConflictError.pkid))
        .group_by(ConflictError.error_type)
        .all()
    )
    type_counts = {row[0]: row[1] for row in type_counts_rows}
    total_conflicts = sum(type_counts.values())

    query = ConflictError.query.order_by(ConflictError.create_dt.desc(), ConflictError.pkid.desc())
    if selected_type and selected_type in ConflictError.ERROR_TYPES:
        query = query.filter(ConflictError.error_type == selected_type)

    conflicts = query.limit(limit).all()

    today = date.today()
    purge_min_date = (today - timedelta(days=365)).isoformat()
    purge_max_date = today.isoformat()

    return render_template(
        "collector/conflicts.html",
        conflicts=conflicts,
        limit=limit,
        selected_type=selected_type,
        type_counts=type_counts,
        total_conflicts=total_conflicts,
        error_types=ConflictError.ERROR_TYPES,
        purge_min_date=purge_min_date,
        purge_max_date=purge_max_date,
    )


@bp.post("/conflicts/<int:pkid>/delete")
@login_required
def delete_conflict(pkid: int):
    conflict = ConflictError.query.get_or_404(pkid)
    db.session.delete(conflict)
    db.session.commit()
    flash("Conflict entry cleared", "success")

    redirect_params: dict[str, str] = {}
    type_param = request.form.get("type") or None
    limit_param = request.form.get("limit") or None
    if type_param:
        redirect_params["type"] = type_param
    if limit_param:
        redirect_params["limit"] = limit_param

    return redirect(url_for("collector.conflicts", **redirect_params))


@bp.post("/conflicts/purge")
@login_required
def purge_conflicts():
    raw_date = request.form.get("purge_date") or ""
    params: dict[str, str] = {}
    type_param = request.form.get("type") or None
    limit_param = request.form.get("limit") or None
    if type_param:
        params["type"] = type_param
    if limit_param:
        params["limit"] = limit_param

    if not raw_date:
        flash("Select a date to purge conflicts.", "warning")
        return redirect(url_for("collector.conflicts", **params))

    try:
        target_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
    except ValueError:
        flash("Invalid purge date format.", "danger")
        return redirect(url_for("collector.conflicts", **params))

    today = date.today()
    min_allowed = today - timedelta(days=365)
    if target_date < min_allowed or target_date > today:
        flash(
            f"Purge date must be between {min_allowed.isoformat()} and {today.isoformat()}.",
            "warning",
        )
        return redirect(url_for("collector.conflicts", **params))

    cutoff = datetime.combine(target_date, datetime.min.time())
    deleted = (
        ConflictError.query
        .filter(ConflictError.create_dt < cutoff)
        .delete(synchronize_session=False)
    )
    db.session.commit()

    if deleted:
        flash(f"Purged {deleted} conflict record(s) before {target_date.isoformat()}.", "success")
    else:
        flash("No conflict records found before the selected date.", "info")

    return redirect(url_for("collector.conflicts", **params))


# -------------------- API: Item search --------------------
@bp.get("/api/items/search")
@login_required
def api_search_items():
    q = (request.args.get("q") or "").strip()
    picker = (request.args.get("picker") or "").strip().lower()
    enforce_company = True
    if picker == "replace":
        enforce_company = False
    elif picker == "item":
        enforce_company = True
    else:
        hx_target = (request.headers.get("HX-Target") or "").strip().lower()
        if "repl-results" in hx_target:
            enforce_company = False

    if not q:
        # For HTMX we still return a table skeleton
        if request.headers.get("HX-Request"):
            return render_template(
                "collector/_item_search_table.html",
                items=[],
                enforce_company_3000=enforce_company,
            )
        return jsonify([])
    limit = min(int(request.args.get("limit", 15) or 15), 50)
    active_only = request.args.get("active_only") == "1"

    query = Item.query
    like_term = f"%{q}%" if len(q) > 3 else f"{q}%"
    query = query.filter(Item.item.ilike(like_term))
    if active_only:
        query = query.filter(Item.is_active == 'Yes')

    items = query.order_by(Item.item).limit(limit).all()

    # Debug logging
    print(f"Search query: '{q}', like_term: '{like_term}', found {len(items)} items")

    if request.headers.get("HX-Request"):
        return render_template(
            "collector/_item_search_table.html",
            items=items,
            enforce_company_3000=enforce_company,
        )

    return jsonify([
        {
            "item": it.item,
            "manufacturer": it.manufacturer,
            "mfg_part_num": it.mfg_part_num,
            "item_description": it.item_description,
            "is_active": it.is_active == 'Yes',
            "is_discontinued": it.is_discontinued == 'Yes',
            "company_3000": it.company_3000 == 'Yes',
        }
        for it in items
    ])


@bp.get("/api/contract-items/search")
@login_required
def api_search_contract_items():
    """Search ContractItem by normalized user input against Item.search_shadow.

    Behavior changes:
      - Only searches on Item.search_shadow (joined via ContractItem.item)
      - Input is normalized by removing spaces and hyphens before building LIKE pattern
      - LIKE pattern: if length > 3 use %term%, else term% (prefix search for short inputs)
    """
    q_raw = (request.args.get("q") or "").strip()
    if not q_raw:
        if request.headers.get("HX-Request"):
            return render_template("collector/_contract_item_search_table.html", contract_items=[])
        return jsonify([])
    # normalize: remove spaces and dashes
    q_norm = q_raw.replace(" ", "").replace("-", "")
    if not q_norm:
        if request.headers.get("HX-Request"):
            return render_template("collector/_contract_item_search_table.html", contract_items=[])
        return jsonify([])
    limit = min(int(request.args.get("limit", 15) or 15), 50)

    query = ContractItem.query
    like_term = f"%{q_norm}%" if len(q_norm) > 3 else f"{q_norm}%"
    query = query.filter(ContractItem.search_shadow.ilike(like_term))

    rows = (
        query.order_by(ContractItem.mfg_part_num, ContractItem.manufacturer, ContractItem.item)
        .limit(limit)
        .all()
    )

    if request.headers.get("HX-Request"):
        return render_template("collector/_contract_item_search_table.html", contract_items=rows)

    return jsonify([
        {
            "contract_id": r.contract_id,
            "manufacturer": r.manufacturer,
            "mfg_part_num": r.mfg_part_num,
            "item_description": r.item_description,
            "item_type": r.item_type,
            "item": r.item,
            "is_mhs": r.is_mhs == 'Yes',
        }
        for r in rows
    ])

# -------------------- API: Batch create links --------------------
@bp.post("/api/item-links/batch")
@login_required
def api_batch_item_links():
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    replace_items = data.get("replace_items") or []
    pending_meta = data.get("pending_meta") or {}
    explicit_stage = data.get("stage") or None
    expected_go_live_date_raw = (data.get("expected_go_live_date") or "").strip() or None

    # Basic validation
    if not items or not replace_items:
        return jsonify({"error": "Both items and replace_items required"}), 400

    from flask import current_app

    max_per_side = current_app.config.get('MAX_BATCH_PER_SIDE', 6)

    try:
        processor = AddItemPairs(
            items=items,
            replace_items=replace_items,
            pending_meta=pending_meta,
            explicit_stage=explicit_stage,
            expected_go_live_date_raw=expected_go_live_date_raw,
            sentinel_replacements=SENTINEL_REPLACEMENTS,
            allowed_stages=ALLOWED_STAGES,
            max_per_side=max_per_side,
            session=db.session,
        )
        result = processor.execute()
    except BatchValidationError as e:
        db.session.rollback()
        payload = {"error": e.message}
        if e.error_code == "missing_items":
            payload["missing"] = e.details.get("missing", [])
        return jsonify(payload), 400
    except ItemGroupConflictError as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 400

    response_body = {
        "group_id": result["records"][0]["item_group"] if result["records"] else None,
        "created": result["created"],
        "skipped": result["skipped"],
        "skipped_details": result.get("skipped_details", []),
        "conflicts": result["conflicts"],
        "stage": result["stage"],
        "stage_locked": result["stage_locked"],
        "merged_groups": result["merged_groups"],
        "records": result["records"],
        "burn_rate_jobs": result.get("burn_rate_jobs", []),
    }
    response_body["group_ids"] = sorted(
        {
            rec.get("item_group")
            for rec in result["records"]
            if isinstance(rec, dict) and rec.get("item_group") is not None
        }
    )
    if result.get("reactivated"):
        response_body["reactivated"] = result["reactivated"]
    status_code = 201 if result["created"] else 200
    return jsonify(response_body), status_code


@bp.post("/api/item-links/upload")
@login_required
def api_upload_item_links():
    file = request.files.get("file")
    filename = file.filename if file else None
    if not file or not filename:
        return jsonify({"status": "error", "error": "Upload a valid Excel file."}), 400

    filename_lower = filename.lower()
    if not filename_lower.endswith(ALLOWED_UPLOAD_EXTENSIONS):
        return jsonify({
            "status": "error",
            "error": "Supported formats: .xlsx, .xlsm, .xltx, .xltm",
        }), 400

    try:
        payload = file.read()
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    except Exception:
        return jsonify({"status": "error", "error": "Unable to read Excel file."}), 400

    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        workbook.close()
        return jsonify({"status": "error", "error": "Excel file is empty."}), 400

    header_map: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        header = str(raw).strip().lower()
        if header:
            header_map[header] = idx

    required_headers = {"item", "replace item"}
    missing_headers = [label for label in required_headers if label not in header_map]
    if missing_headers:
        workbook.close()
        return jsonify({
            "status": "error",
            "error": f"Missing required column(s): {', '.join(missing_headers)}",
        }), 400

    item_col = header_map["item"]
    replace_col = header_map["replace item"]
    max_col = max(item_col, replace_col) + 1

    def _coerce_cell(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            if isinstance(value, float):
                if not value.is_integer():
                    return str(value).strip()
                value = int(value)
            return str(value)
        text = str(value).strip()
        if not text:
            return ""
        if "." in text:
            try:
                numeric = float(text)
            except ValueError:
                return text
            if numeric.is_integer():
                return str(int(numeric))
        return text

    def _parse_code(label: str, text: str, *, allow_sentinel: bool = False) -> tuple[str | None, list[str]]:
        if not text:
            return None, [f"{label} is required."]
        upper = text.upper()
        if allow_sentinel and upper == "NO REPLACEMENT":
            return "NO REPLACEMENT", []
        if not text.isdigit():
            return None, [f"{label} must be a six-digit number."]
        if len(text) != 6:
            return None, [f"{label} must be a six-digit number."]
        if text.startswith("0"):
            return None, [f"{label} cannot start with 0."]
        if not ITEM_CODE_PATTERN.match(text):
            return None, [f"{label} must be a six-digit number."]
        return text, []

    parsed_rows: list[dict[str, object]] = []
    data_rows = sheet.iter_rows(min_row=2, min_col=1, max_col=max_col, values_only=True)
    for row_index, row in enumerate(data_rows, start=2):
        row = row or ()
        item_text = _coerce_cell(row[item_col] if item_col < len(row) else None)
        replace_text = _coerce_cell(row[replace_col] if replace_col < len(row) else None)
        if not item_text and not replace_text:
            continue
        row_info: dict[str, object] = {
            "row_number": row_index,
            "item_input": item_text,
            "replace_input": replace_text,
            "errors": [],
        }
        item_code, item_errors = _parse_code("Item", item_text)
        if item_errors:
            row_info["errors"].extend(item_errors)
        row_info["item_code"] = item_code
        replace_code, replace_errors = _parse_code("Replace Item", replace_text, allow_sentinel=True)
        if replace_errors:
            row_info["errors"].extend(replace_errors)
        row_info["replace_code"] = replace_code
        row_info["replace_is_sentinel"] = replace_code == "NO REPLACEMENT"
        parsed_rows.append(row_info)

    workbook.close()

    if not parsed_rows:
        return jsonify({"status": "error", "error": "No data rows found in Excel file."}), 400

    codes_to_lookup = {
        row["item_code"]
        for row in parsed_rows
        if not row["errors"] and row.get("item_code")
    }
    codes_to_lookup.update(
        {
            row["replace_code"]
            for row in parsed_rows
            if not row["errors"]
            and row.get("replace_code")
            and not row.get("replace_is_sentinel")
        }
    )

    item_lookup: dict[str, Item] = {}
    if codes_to_lookup:
        items = Item.query.filter(Item.item.in_(codes_to_lookup)).all()
        item_lookup = {item.item: item for item in items}

    for row in parsed_rows:
        if row["errors"]:
            continue
        item_code = row["item_code"]
        item_record = item_lookup.get(item_code)
        if not item_record:
            row["errors"].append(f"Item {item_code} not found in PLM items view (check on Infor to confirm the item is valid).")
            continue
        if (item_record.company_3000 or "").strip() != "Yes":
            row["errors"].append(f"Item {item_code} is not available in company 3000.")
        if row.get("replace_is_sentinel"):
            continue
        repl_code = row["replace_code"]
        repl_record = item_lookup.get(repl_code)
        if not repl_record:
            row["errors"].append(f"Replace Item {repl_code} not found in PLM items view (check on Infor to confirm the item is valid).")
            continue

    from flask import current_app

    max_per_side = current_app.config.get("MAX_BATCH_PER_SIDE", 6)

    total_created = 0
    total_reactivated = 0
    all_skipped: list[list[object]] = []
    all_skipped_details: list[dict[str, object]] = []
    all_conflicts: list[dict[str, object]] = []
    all_records: list[dict[str, object]] = []
    burn_rate_jobs: list[dict[str, object]] = []

    candidate_rows = [row for row in parsed_rows if not row["errors"]]
    for row in candidate_rows:
        try:
            processor = AddItemPairs(
                items=[row["item_code"]],
                replace_items=[row["replace_code"]],
                pending_meta={},
                explicit_stage=None,
                expected_go_live_date_raw=None,
                sentinel_replacements=SENTINEL_REPLACEMENTS,
                allowed_stages=ALLOWED_STAGES,
                max_per_side=max_per_side,
                session=db.session,
            )
            result = processor.execute()
        except BatchValidationError as exc:
            db.session.rollback()
            row["errors"].append(exc.message)
            continue
        except ItemGroupConflictError as exc:
            db.session.rollback()
            row["errors"].append(str(exc))
            continue
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception(
                "Batch upload ingestion failed for row %s", row.get("row_number")
            )
            row["errors"].append("Unexpected error while processing row.")
            detail = str(exc).strip()
            if detail:
                row["errors"].append(detail)
            continue

        total_created += result.get("created", 0)
        if result.get("reactivated"):
            total_reactivated += result["reactivated"]
        skipped_rows = result.get("skipped", []) or []
        if skipped_rows:
            all_skipped.extend(skipped_rows)
        skipped_details = result.get("skipped_details", []) or []
        if skipped_details:
            all_skipped_details.extend(skipped_details)
        conflicts = result.get("conflicts", []) or []
        if conflicts:
            all_conflicts.extend(conflicts)
        records = result.get("records", []) or []
        if records:
            all_records.extend(records)
        jobs = result.get("burn_rate_jobs", []) or []
        if jobs:
            burn_rate_jobs.extend(jobs)

    def _serialize_invalid_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
        payload: list[dict[str, object]] = []
        for row in rows:
            errors = row.get("errors") or []
            if not errors:
                continue
            item_display = row.get("item_code") or row.get("item_input") or ""
            replace_display = "NO REPLACEMENT"
            if not row.get("replace_is_sentinel"):
                replace_display = row.get("replace_code") or row.get("replace_input") or ""
            payload.append(
                {
                    "row_number": row.get("row_number"),
                    "item": item_display,
                    "replace_item": replace_display,
                    "messages": errors,
                }
            )
        return payload

    invalid_rows = _serialize_invalid_rows(parsed_rows)
    successful_rows = [row for row in parsed_rows if not row.get("errors")]

    # Deduplicate burn-rate jobs by job_id to avoid duplicates across rows
    deduped_jobs: list[dict[str, object]] = []
    seen_job_ids: set[int] = set()
    for job in burn_rate_jobs:
        job_id = job.get("job_id")
        if job_id is not None:
            if job_id in seen_job_ids:
                continue
            seen_job_ids.add(job_id)
        deduped_jobs.append(job)

    group_ids = sorted({rec.get("item_group") for rec in all_records if rec.get("item_group") is not None})

    response_body = {
        "status": "ok" if successful_rows else "error",
        "submitted_rows": len(parsed_rows),
        "processed_rows": len(successful_rows),
        "created": total_created,
        "reactivated": total_reactivated,
        "skipped": all_skipped,
        "skipped_details": all_skipped_details,
        "conflicts": all_conflicts,
        "records": all_records,
        "burn_rate_jobs": deduped_jobs,
        "invalid_rows": invalid_rows,
        "group_ids": group_ids,
    }

    if successful_rows:
        response_body["message"] = "Batch upload processed."
        status_code = 200
    else:
        response_body["error"] = "No rows were processed successfully."
        status_code = 400

    return jsonify(response_body), status_code


@bp.get("/api/burn-rate-jobs")
@login_required
def api_burn_rate_jobs():
    def _parse_int_params(raw: str | None) -> list[int]:
        if not raw:
            return []
        values: list[int] = []
        for part in raw.split(","):
            part = part.strip()
            if not part:
                continue
            try:
                values.append(int(part))
            except ValueError:
                continue
        return values

    job_ids = _parse_int_params(request.args.get("job_ids"))
    item_link_ids = _parse_int_params(request.args.get("item_link_ids"))

    query = BurnRateRefreshJob.query
    if job_ids:
        query = query.filter(BurnRateRefreshJob.id.in_(job_ids))
    elif item_link_ids:
        query = query.filter(BurnRateRefreshJob.item_link_id.in_(item_link_ids))

    rows = (
        query.order_by(BurnRateRefreshJob.created_at.desc())
        .limit(100)
        .all()
    )

    payload = [
        {
            "job_id": row.id,
            "item_link_id": row.item_link_id,
            "status": row.status,
            "message": row.message,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "started_at": row.started_at.isoformat() if row.started_at else None,
            "finished_at": row.finished_at.isoformat() if row.finished_at else None,
        }
        for row in rows
    ]

    return jsonify(payload)


# -------------------- API: Delete single item link --------------------
@bp.delete('/api/item-links/<item>/<replace_item>')
@login_required
def api_delete_item_link(item: str, replace_item: str):
    # Normalize textual null markers coming from client URLs to None so
    # we can delete rows where replace_item IS NULL.
    if isinstance(replace_item, str) and replace_item.lower() in ('none', 'null', 'nan', ''):
        replace_item = None
    record = ItemLink.query.filter_by(item=item, replace_item=replace_item).first()
    if not record:
        return jsonify({"error": "Not found"}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({"status": "deleted", "item": item, "replace_item": replace_item})
