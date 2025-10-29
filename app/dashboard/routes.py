import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from flask import Blueprint, render_template, request, jsonify, send_file, abort
from flask_login import login_required
from sqlalchemy import select, func
from ..utility.item_locations import build_location_pairs
from .. import db
from ..models.inventory import Requesters365Day
from ..models.relations import ItemLink, PLMTrackerBase, PLMQty, PLMDailyIssueOutQty
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

bp = Blueprint("dashboard", __name__, url_prefix="/dashboard")


TRI_STATE_VALUES = {"yes", "no", "blank"}
ALLOWED_STAGE_VALUES = {
    "Tracking - Discontinued",
    "Tracking - Item Transition",
    "Pending Clinical Readiness",
}


@bp.route("/documents/order-point-calculation")
@login_required
def order_point_calc_doc():
    return render_template("documents/orderPointCalc.html")


def _normalize_code(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _collect_item_pool(rows: list[dict]) -> set[str]:
    items: set[str] = set()
    for row in rows:
        item = _normalize_code(row.get("item"))
        if item:
            items.add(item)
    return items


def _aggregate_requester_rows(raw_rows: list[dict]) -> list[dict]:
    aggregated: dict[str, dict] = {}
    for row in raw_rows:
        requester = _normalize_code(row.get("requester"))
        if not requester:
            continue
        entry = aggregated.setdefault(
            requester,
            {
                "requester": requester,
                "name": "",
                "email": "",
                "locations": set(),
                "items": set(),
                "requisition_ids": set(),
                "request_count": 0,
            },
        )

        name = _normalize_code(row.get("name"))
        if name and not entry["name"]:
            entry["name"] = name

        email = _normalize_code(row.get("email"))
        if email and not entry["email"]:
            entry["email"] = email

        location = _normalize_code(row.get("location"))
        if location:
            entry["locations"].add(location)

        item = _normalize_code(row.get("item"))
        if item:
            entry["items"].add(item)

        requisition = _normalize_code(row.get("requisition"))
        if requisition:
            entry["requisition_ids"].add(requisition)

        count_value = row.get("requests_count")
        if count_value is None:
            count_value = row.get("RequestsCount")
        try:
            parsed_count = int(count_value) if count_value is not None else 0
        except (TypeError, ValueError):
            parsed_count = 0

        if parsed_count < 0:
            parsed_count = 0

        entry["request_count"] += parsed_count

    results: list[dict] = []
    for entry in aggregated.values():
        results.append(
            {
                "requester": entry["requester"],
                "name": entry["name"] or None,
                "email": entry["email"] or None,
                "locations": sorted(entry["locations"]),
                "items": sorted(entry["items"]),
                "requisition_ids": sorted(entry["requisition_ids"]),
                "request_count": entry["request_count"],
            }
        )

    def _sort_key(payload: dict) -> tuple[str, str]:
        name_key = (payload.get("name") or "").lower()
        return (name_key, payload.get("requester") or "")

    results.sort(key=_sort_key)
    return results


def _normalize_tri_state(value: object) -> str:
    """Normalize database values to 'yes', 'no', or 'blank'."""
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (int, float)):
        if value == 0:
            return "no"
        if value == 1:
            return "yes"
    text = str(value).strip().lower()
    if text in ("", "na", "n/a", "none", "null"):
        return "blank"
    if text in ("yes", "y", "true", "t", "1", "active"):
        return "yes"
    if text in ("no", "n", "false", "f", "0", "inactive"):
        return "no"
    return "blank"


def _apply_tri_state_filter(rows, key: str, desired: str | None):
    target = (desired or "").strip().lower()
    if target not in TRI_STATE_VALUES:
        return rows
    return [row for row in rows if _normalize_tri_state(row.get(key)) == target]


def _parse_stage_values(args) -> list[str]:
    stage_param = args.get("stage") or args.get("stages")
    if stage_param:
        stage_list_raw = [s.strip() for s in stage_param.split(",") if s.strip()]
        stages = [s for s in stage_list_raw if s in ALLOWED_STAGE_VALUES]
        return stages or list(ALLOWED_STAGE_VALUES)
    return list(ALLOWED_STAGE_VALUES)


def _parse_item_group_filters(param: str | None) -> list[int]:
    if not param:
        return []
    filters: list[int] = []
    for part in param.split(','):
        part = part.strip()
        if not part:
            continue
        try:
            filters.append(int(part))
        except ValueError:
            continue
    # dedupe preserving order
    seen = set()
    out: list[int] = []
    for value in filters:
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def _parse_location_filters(param: str | None) -> list[str]:
    if not param:
        return []
    return [loc.strip() for loc in param.split(',') if loc.strip()]


def _desc_search_lower(args) -> str:
    return (args.get("desc_search") or "").lower().strip()


def _normalize_text(val: object) -> str:
    return str(val or "").strip().lower()


def _is_r_only_location(row: dict) -> bool:
    if not isinstance(row, dict):
        return False
    loc = _normalize_text(row.get("location"))
    loc_type = _normalize_text(row.get("location_type"))
    if not loc:
        return True
    if "r-only" in loc_type:
        return True
    compact = " ".join(loc.split())
    return compact in {"r-only", "r only", "r-only location", "r only location"}


def _coerce_excel_value(value):
    if value is None:
        return ""
    return value


def _to_decimal(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    try:
        return Decimal(text.replace(',', ''))
    except (InvalidOperation, ValueError):
        return None


def _apply_quantity_filter(rows, field: str, desired: str | None):
    target = (desired or "").strip().lower()
    if target not in {"zero", "positive"}:
        return rows
    filtered: list[dict] = []
    for row in rows:
        val = _to_decimal(row.get(field))
        if val is None:
            continue
        if target == "zero" and val == 0:
            filtered.append(row)
        elif target == "positive" and val > 0:
            filtered.append(row)
    return filtered


def _filtered_inventory_rows(args, *, apply_filters: bool = True) -> list[dict]:
    if apply_filters:
        stages_list = _parse_stage_values(args)
        item_group_filters = _parse_item_group_filters(args.get("item_group"))
        location_filters = _parse_location_filters(args.get("location"))
        company = args.get("company") or None
        active_param = args.get("active")
        require_active = active_param.lower() == "true" if active_param else False
        desc_search_lower = _desc_search_lower(args)
    else:
        stages_list = list(ALLOWED_STAGE_VALUES)
        item_group_filters: list[int] = []
        location_filters: list[str] = []
        company = None
        require_active = False
        desc_search_lower = ""

    all_rows = build_location_pairs(
        stages=stages_list,
        company=company,
        location=None,
        require_active=require_active,
        include_par=False,
        location_types=["Inventory Location"],
    )
    if location_filters:
        loc_set = set(location_filters)
        all_rows = [r for r in all_rows if (r.get("group_location") in loc_set)]
    if item_group_filters:
        allowed_set = set(item_group_filters)
        all_rows = [r for r in all_rows if r.get("item_group") in allowed_set]
    if desc_search_lower:
        all_rows = [
            r
            for r in all_rows
            if (
                str(r.get("item_description") or "").lower().find(desc_search_lower) != -1
                or str(r.get("item_description_ri") or "").lower().find(desc_search_lower) != -1
            )
        ]

    if apply_filters:
        all_rows = _apply_tri_state_filter(all_rows, "auto_replenishment", args.get("auto_repl_state"))
        all_rows = _apply_tri_state_filter(all_rows, "active", args.get("active_state"))
        all_rows = _apply_tri_state_filter(all_rows, "discontinued", args.get("discontinued_state"))
        all_rows = _apply_tri_state_filter(all_rows, "auto_replenishment_ri", args.get("auto_repl_state_ri"))
        all_rows = _apply_tri_state_filter(all_rows, "active_ri", args.get("active_state_ri"))
        all_rows = _apply_tri_state_filter(all_rows, "discontinued_ri", args.get("discontinued_state_ri"))
        all_rows = _apply_quantity_filter(all_rows, "current_qty", args.get("current_qty_filter"))
        all_rows = _apply_quantity_filter(all_rows, "current_qty_ri", args.get("current_qty_ri_filter"))
    return all_rows


def _filtered_par_rows(args, *, apply_filters: bool = True) -> list[dict]:
    if apply_filters:
        stages_list = _parse_stage_values(args)
        item_group_filters = _parse_item_group_filters(args.get("item_group"))
        location_filters = _parse_location_filters(args.get("location"))
        desc_search_lower = _desc_search_lower(args)
    else:
        stages_list = list(ALLOWED_STAGE_VALUES)
        item_group_filters = []
        location_filters = []
        desc_search_lower = ""

    all_rows = build_location_pairs(
        stages=stages_list,
        location=None,
        include_par=True,
        location_types=["Par Location"],
    )
    if location_filters:
        loc_set = set(location_filters)
        all_rows = [r for r in all_rows if (r.get("group_location") in loc_set)]
    if item_group_filters:
        allowed_set = set(item_group_filters)
        all_rows = [r for r in all_rows if r.get("item_group") in allowed_set]
    if desc_search_lower:
        all_rows = [
            r
            for r in all_rows
            if (
                str(r.get("item_description") or "").lower().find(desc_search_lower) != -1
                or str(r.get("item_description_ri") or "").lower().find(desc_search_lower) != -1
            )
        ]

    if apply_filters:
        all_rows = _apply_tri_state_filter(all_rows, "auto_replenishment", args.get("auto_repl_state"))
        all_rows = _apply_tri_state_filter(all_rows, "active", args.get("active_state"))
        all_rows = _apply_tri_state_filter(all_rows, "discontinued", args.get("discontinued_state"))
        all_rows = _apply_tri_state_filter(all_rows, "auto_replenishment_ri", args.get("auto_repl_state_ri"))
        all_rows = _apply_tri_state_filter(all_rows, "active_ri", args.get("active_state_ri"))
        all_rows = _apply_tri_state_filter(all_rows, "discontinued_ri", args.get("discontinued_state_ri"))

    for r in all_rows:
        reorder_pt = r.get("reorder_point")
        weekly_burn = r.get("weekly_burn")
        try:
            if reorder_pt is None or weekly_burn in (None, 0, "0", "0.0"):
                r["weeks_reorder"] = "unknown"
            else:
                val = float(reorder_pt) / float(weekly_burn) if float(weekly_burn) != 0 else None
                r["weeks_reorder"] = "unknown" if val is None else val
        except Exception:
            r["weeks_reorder"] = "unknown"
        reorder_pt_ri = r.get("reorder_point_ri")
        weekly_burn_ri = r.get("weekly_burn_ri")
        try:
            if reorder_pt_ri is None or weekly_burn_ri in (None, 0, "0", "0.0"):
                r["weeks_reorder_ri"] = "unknown"
            else:
                val = float(reorder_pt_ri) / float(weekly_burn_ri) if float(weekly_burn_ri) != 0 else None
                r["weeks_reorder_ri"] = "unknown" if val is None else val
        except Exception:
            r["weeks_reorder_ri"] = "unknown"
    return all_rows


INVENTORY_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Stage", "stage"),
    ("Item Group", "item_group"),
    ("Group Type", "group_type"),
    ("Weekly Burn (G. & Loc.)", "weekly_burn_group_location"),
    ("Item", "item"),
    ("Location", "location"),
    ("Location Text", "location_text"),
    ("Company", "company"),
    ("Preferred Bin", "preferred_bin"),
    ("Auto-repl.", "auto_replenishment"),
    ("Active", "active"),
    ("Discon.", "discontinued"),
    ("Current Qty", "current_qty"),
    ("Weekly Burn", "weekly_burn"),
    ("Weeks on Hand", "weeks_on_hand"),
    ("Stock UOM", "stock_uom"),
    ("UOM Conversion", "uom_conversion"),
    ("Buy UOM", "buy_uom"),
    ("Buy UOM Multiplier", "buy_uom_multiplier"),
    ("Transaction UOM", "transaction_uom"),
    ("Transaction UOM Multiplier", "transaction_uom_multiplier"),
    ("Reorder Policy", "reorder_quantity_code"),
    ("Reorder Point", "reorder_point"),
    ("Min Order Qty", "min_order_qty"),
    ("Max Order Qty", "max_order_qty"),
    ("Manufacturer Number", "manufacturer_number"),
    ("90-day PO Qty", "po_90_qty"),
    ("Repl. Item", "replacement_item"),
    ("Location (RI)", "location_ri"),
    ("Location Text (RI)", "location_text_ri"),
    ("Company (RI)", "company_ri"),
    ("Preferred Bin (RI)", "preferred_bin_ri"),
    ("Preferred Bin (Recom.)", "recommended_preferred_bin_ri"),
    ("Auto-repl. (RI)", "auto_replenishment_ri"),
    ("Auto-repl. (Recom.)", "recommended_auto_replenishment_ri"),
    ("Active (RI)", "active_ri"),
    ("Discon. (RI)", "discontinued_ri"),
    ("Current Qty (RI)", "current_qty_ri"),
    ("Weekly Burn (RI)", "weekly_burn_ri"),
    ("Weeks on Hand (RI)", "weeks_on_hand_ri"),
    ("Stock UOM (RI)", "stock_uom_ri"),
    ("UOM Conversion (RI)", "uom_conversion_ri"),
    ("Buy UOM (RI)", "buy_uom_ri"),
    ("Buy UOM Multiplier (RI)", "buy_uom_multiplier_ri"),
    ("Transaction UOM (RI)", "transaction_uom_ri"),
    ("Transaction UOM (Recom.)", "recommended_transaction_uom_ri"),
    ("Transaction UOM Multiplier (RI)", "transaction_uom_multiplier_ri"),
    ("Transaction UOM Multiplier (Recom.)", "recommended_transaction_uom_multiplier_ri"),
    ("Reorder Policy (RI)", "reorder_quantity_code_ri"),
    ("Reorder Policy (Recom.)", "recommended_reorder_quantity_code_ri"),
    ("Reorder Point (RI)", "reorder_point_ri"),
    ("Reorder Point (Recom.)", "recommended_reorder_point_ri"),
    ("Min Order Qty (RI)", "min_order_qty_ri"),
    ("Min Order Qty (Recom.)", "recommended_min_order_qty_ri"),
    ("Max Order Qty (RI)", "max_order_qty_ri"),
    ("Max Order Qty (Recom.)", "recommended_max_order_qty_ri"),
    ("Manufacturer Number (RI)", "manufacturer_number_ri"),
    ("Item Description", "item_description"),
    ("Item Description (RI)", "item_description_ri"),
    ("Setup Action", "action"),
    ("Notes", "notes"),
]


PAR_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Stage", "stage"),
    ("Item Group", "item_group"),
    ("Group Type", "group_type"),
    ("Weekly Burn (G. & Loc.)", "weekly_burn_group_location"),
    ("Item", "item"),
    ("Location", "location"),
    ("Location Text", "location_text"),
    ("Company", "company"),
    ("Preferred Bin", "preferred_bin"),
    ("Auto-repl.", "auto_replenishment"),
    ("Active", "active"),
    ("Discon.", "discontinued"),
    ("Reorder Point", "reorder_point"),
    ("Weekly Demand", "weekly_burn"),
    ("Weeks Reorder", "weeks_reorder"),
    ("Stock UOM", "stock_uom"),
    ("UOM Conversion", "uom_conversion"),
    ("Buy UOM", "buy_uom"),
    ("Buy UOM Multiplier", "buy_uom_multiplier"),
    ("Transaction UOM", "transaction_uom"),
    ("Transaction UOM Multiplier", "transaction_uom_multiplier"),
    ("Reorder Policy", "reorder_quantity_code"),
    ("Min Order Qty", "min_order_qty"),
    ("Max Order Qty", "max_order_qty"),
    ("Manufacturer Number", "manufacturer_number"),
    ("90-day Req Qty", "req_qty_ea"),
    ("Repl. Item", "replacement_item"),
    ("Location (RI)", "location_ri"),
    ("Location Text (RI)", "location_text_ri"),
    ("Company (RI)", "company_ri"),
    ("Preferred Bin (RI)", "preferred_bin_ri"),
    ("Preferred Bin (Recom.)", "recommended_preferred_bin_ri"),
    ("Auto-repl. (RI)", "auto_replenishment_ri"),
    ("Auto-repl. (Recom.)", "recommended_auto_replenishment_ri"),
    ("Active (RI)", "active_ri"),
    ("Discon. (RI)", "discontinued_ri"),
    ("Reorder Point (RI)", "reorder_point_ri"),
    ("Reorder Point (Recom.)", "recommended_reorder_point_ri"),
    ("Weekly Demand (RI)", "weekly_burn_ri"),
    ("Weeks Reorder (RI)", "weeks_reorder_ri"),
    ("Stock UOM (RI)", "stock_uom_ri"),
    ("UOM Conversion (RI)", "uom_conversion_ri"),
    ("Buy UOM (RI)", "buy_uom_ri"),
    ("Buy UOM Multiplier (RI)", "buy_uom_multiplier_ri"),
    ("Transaction UOM (RI)", "transaction_uom_ri"),
    ("Transaction UOM (Recom.)", "recommended_transaction_uom_ri"),
    ("Transaction UOM Multiplier (RI)", "transaction_uom_multiplier_ri"),
    ("Transaction UOM Multiplier (Recom.)", "recommended_transaction_uom_multiplier_ri"),
    ("Reorder Policy (RI)", "reorder_quantity_code_ri"),
    ("Reorder Policy (Recom.)", "recommended_reorder_quantity_code_ri"),
    ("Min Order Qty (RI)", "min_order_qty_ri"),
    ("Min Order Qty (Recom.)", "recommended_min_order_qty_ri"),
    ("Max Order Qty (RI)", "max_order_qty_ri"),
    ("Max Order Qty (Recom.)", "recommended_max_order_qty_ri"),
    ("Manufacturer Number (RI)", "manufacturer_number_ri"),
    ("Item Description", "item_description"),
    ("Item Description (RI)", "item_description_ri"),
    ("Setup Action", "action"),
    ("Notes", "notes"),
]


CUSTOM_EXPORT_MODES: set[str] = {"custom", "inventory_setup", "par_setup"}

INVENTORY_SETUP_HEADER_OVERRIDES: dict[str, str] = {
    "company": "Company",
    "location_ri": "InventoryLocation",
    "group_type": "Group Type",
    "replacement_item": "Item",
    "item": " Original Item",
    "recommended_transaction_uom_ri": "DefaultTransactionUOM-Issue UOM",
    "recommended_preferred_bin_ri": "PreferredBin",
    "recommended_min_order_qty_ri": "MinimumOrderQuantity",
    "recommended_max_order_qty_ri": "MaximumOrderQuantity",
    "recommended_reorder_point_ri": "ReorderPoint",
    "recommended_auto_replenishment_ri": "AutomaticPurchaseOrder (True/Falsse)",
    "manufacturer_number_ri": "Item Manufacturer Number",
    "action": "Requested update/Action",
    "notes": "Notes",
    "preferred_bin_ri": "Current PreferredBin (Repl. Item)",
    "min_order_qty_ri": "Current Min (Repl. Item)",
    "max_order_qty_ri": "Current Max (Repl. Item)",
    "reorder_point_ri": "Current Reorder (Repl. Item)",
}

PAR_SETUP_HEADER_OVERRIDES: dict[str, str] = {
    "company": "Company",
    "location_ri": "Inventory Location",
    "location_text": "Inventory Location Name",
    "group_type": "Group Type",
    "replacement_item": "Item",
    "item": " Original Item",
    "manufacturer_number_ri": "Item Manufacturer Number",
    "item_description_ri": "Item.Description",
    "recommended_min_order_qty_ri": "Min",
    "recommended_max_order_qty_ri": "Max",
    "recommended_reorder_point_ri": "ReorderPoint",
    "stock_uom_ri": "UOM Unit Of Measure",
    "recommended_preferred_bin_ri": "BIN Location                        (All New Sequence)",
    "action": "Requested update/Action",
    "notes": "Notes",
    "preferred_bin_ri": "Current Bin (Repl. Item)",
    "reorder_point_ri": "Current Reorder (Repl. Item)",
}

PRESET_HEADER_OVERRIDES: dict[str, dict[str, str]] = {
    "inventory_setup": INVENTORY_SETUP_HEADER_OVERRIDES,
    "par_setup": PAR_SETUP_HEADER_OVERRIDES,
}


def _parse_column_selection(param: str | None) -> list[str]:
    if not param:
        return []
    seen: set[str] = set()
    results: list[str] = []
    for part in param.split(","):
        field = part.strip()
        if not field or field in seen:
            continue
        seen.add(field)
        results.append(field)
    return results


def _filter_export_columns(
    column_defs: list[tuple[str, str]],
    requested_fields: list[str],
) -> list[tuple[str, str]]:
    if not requested_fields:
        return []
    lookup = {field_name: (header, field_name) for header, field_name in column_defs}
    filtered: list[tuple[str, str]] = []
    for field in requested_fields:
        column = lookup.get(field)
        if column and column not in filtered:
            filtered.append(column)
    return filtered


@bp.route("/")
@login_required
def index():
    return render_template("dashboard/index.html")

@bp.route("/groups/<int:group_id>")
@login_required
def group_detail(group_id: int):
    # Placeholder; later will query KPIs
    return render_template("dashboard/group.html", group_id=group_id)


@bp.route("/api/inventory")
@login_required
def api_inventory():
    # Pagination params
    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get("per_page", 20))
    except ValueError:
        per_page = 20
    page = max(page, 1)
    per_page = max(min(per_page, 200), 1)

    all_rows = _filtered_inventory_rows(request.args)
    total = len(all_rows)
    start = (page - 1) * per_page
    end = start + per_page
    rows = all_rows[start:end]
    if rows:
        print(rows[0]) #debug (keep it for now)
    return jsonify({
        "rows": rows,
        "count": len(rows),
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if per_page else 1,
    })


@bp.route("/api/filter-options")
@login_required
def api_filter_options():
    """Return distinct lists for dashboard filters.

    - item_groups: from ItemLink.item_group (exclude NULL, sorted ascending)
      with associated item numbers formatted as {group_id: [item1, item2, ...]}
    - locations: distinct (LocationType, Location) from PLMTrackerBase limited to inventory location types
      formatted as "{LocationType} - {Location}" and include raw location for querying.
    - stages: allowed stage values (static list)
    """
    # Item Groups with associated items
    from ..models.relations import ItemGroup
    
    # Get all item groups first
    item_groups_query = select(func.distinct(ItemLink.item_group)).where(ItemLink.item_group.isnot(None)).order_by(ItemLink.item_group)
    group_ids = [row[0] for row in db.session.execute(item_groups_query).all()]
    
    # Build item groups data structure: [{value: group_id, items: [item1, item2, ...], label: "123 - item1, item2, item3"}]
    item_groups = []
    for group_id in group_ids:
        # Query ItemGroup table to get all distinct items for this group
        items_query = (
            select(func.distinct(ItemGroup.item))
            .where(ItemGroup.item_group == group_id)
            .where(ItemGroup.item.isnot(None))
            .order_by(ItemGroup.item)
        )
        items = [row[0] for row in db.session.execute(items_query).all()]
        
        # Format label as "Group ID - item1, item2, item3"
        items_str = ", ".join(items) if items else ""
        label = f"{group_id} - {items_str}" if items_str else str(group_id)
        
        item_groups.append({
            "value": group_id,
            "items": items,
            "label": label
        })

    # Locations (pull from view)
    v = PLMTrackerBase
    loc_query = (
        select(func.distinct(v.LocationType), v.Group_Locations)
        .where(v.Group_Locations.isnot(None))
        .order_by(v.LocationType, v.Group_Locations)
    )
    locations = []
    for lt, group_loc in db.session.execute(loc_query).all():
        label = f"{lt} - {group_loc}" if lt else group_loc
        locations.append({"value": group_loc, "type": lt, "label": label})

    stages = [
        "Tracking - Discontinued",
        "Tracking - Item Transition",
        "Pending Clinical Readiness",
    ]

    return jsonify({
        "item_groups": item_groups,
        "locations": locations,
        "stages": stages,
    })


@bp.route("/api/par")
@login_required
def api_par():
    """Par location data (Par Locations table).

    Similar filtering semantics as inventory endpoint but restricted to Par Location types.
    Weeks Reorder = ReorderPoint / weekly_burn (and we negate to present positive like inventory logic).
    """
    # Pagination
    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get("per_page", 100))
    except ValueError:
        per_page = 100
    page = max(page, 1)
    per_page = max(min(per_page, 200), 1)

    all_rows = _filtered_par_rows(request.args)
    total = len(all_rows)
    start = (page - 1) * per_page
    end = start + per_page
    rows = all_rows[start:end]
    return jsonify({
        "rows": rows,
        "count": len(rows),
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if per_page else 1,
    })


@bp.route("/api/requesters")
@login_required
def api_requesters():
    inventory_rows = _filtered_inventory_rows(request.args)
    par_rows = _filtered_par_rows(request.args)

    hide_r_only = (request.args.get("hide_r_only") or "").strip().lower() == "true"
    if hide_r_only:
        inventory_rows = [row for row in inventory_rows if not _is_r_only_location(row)]
        par_rows = [row for row in par_rows if not _is_r_only_location(row)]

    item_pool = _collect_item_pool(inventory_rows) | _collect_item_pool(par_rows)
    if not item_pool:
        return jsonify({
            "items": [],
            "requesters": [],
            "requester_count": 0,
            "email_addresses": [],
        })

    stmt = (
        select(
            Requesters365Day.Requester.label("requester"),
            Requesters365Day.RequesterName.label("name"),
            Requesters365Day.EmailAddress.label("email"),
            Requesters365Day.RequestingLocation.label("location"),
            Requesters365Day.Item.label("item"),
            Requesters365Day.Requisition_FD5.label("requisition"),
            Requesters365Day.RequestsCount.label("requests_count"),
        )
        .where(Requesters365Day.Item.in_(sorted(item_pool)))
        .where(Requesters365Day.RequestingLocation.like("R%"))
    )

    requester_rows = [dict(row._mapping) for row in db.session.execute(stmt)]
    requesters = _aggregate_requester_rows(requester_rows)
    email_addresses = sorted({r["email"] for r in requesters if r["email"]})

    return jsonify({
        "items": sorted(item_pool),
        "requesters": requesters,
        "requester_count": len(requesters),
        "email_addresses": email_addresses,
    })


@bp.route("/api/refresh-timestamp")
@login_required
def api_refresh_timestamp():
    """Return the latest successful data refresh timestamp from process log."""
    refresh_timestamp = None
    try:
        from ..models.log import ProcessLog
        latest_refresh = ProcessLog.get_latest_success_timestamp(db.session)
        print(f"[DEBUG] Latest refresh from DB: {latest_refresh}")
        if latest_refresh:
            refresh_timestamp = latest_refresh.isoformat()
            print(f"[DEBUG] Formatted timestamp: {refresh_timestamp}")
    except Exception as e:
        print(f"[ERROR] Failed to get refresh timestamp: {e}")
        import traceback
        traceback.print_exc()
    
    return jsonify({
        "refresh_timestamp": refresh_timestamp,
    })


@bp.route("/api/stats")
@login_required
def api_stats():
    """Return KPI statistics based on applied filters.

    Returns:
    - distinct_groups: count of distinct item_group values
    - distinct_items: count of distinct items (including replacement_item)
    - distinct_locations: count of distinct locations (from both inventory and par)
    """
    inventory_rows = _filtered_inventory_rows(request.args)
    par_rows = _filtered_par_rows(request.args)

    # Collect distinct item groups
    groups_set = set()
    for row in inventory_rows:
        if row.get("item_group"):
            groups_set.add(row.get("item_group"))
    for row in par_rows:
        if row.get("item_group"):
            groups_set.add(row.get("item_group"))

    # Collect distinct items (including replacement items)
    items_set = set()
    for row in inventory_rows:
        if row.get("item"):
            items_set.add(row.get("item"))
        if row.get("replacement_item"):
            items_set.add(row.get("replacement_item"))
    for row in par_rows:
        if row.get("item"):
            items_set.add(row.get("item"))
        if row.get("replacement_item"):
            items_set.add(row.get("replacement_item"))

    # Collect distinct locations (using group_location as the canonical location identifier)
    locations_set = set()
    for row in inventory_rows:
        loc = row.get("group_location") or row.get("location")
        if loc:
            locations_set.add(loc)
    for row in par_rows:
        loc = row.get("group_location") or row.get("location")
        if loc:
            locations_set.add(loc)

    return jsonify({
        "distinct_groups": len(groups_set),
        "distinct_items": len(items_set),
        "distinct_locations": len(locations_set),
    })


@bp.route("/export/<string:table_key>")
@login_required
def export_table(table_key: str):
    table_key_normalized = table_key.lower()
    row_scope = (request.args.get("row_scope") or "filtered").strip().lower()
    if row_scope not in {"all", "filtered"}:
        row_scope = "filtered"
    apply_filters = row_scope != "all"

    if table_key_normalized == "inventory":
        rows = _filtered_inventory_rows(request.args, apply_filters=apply_filters)
        columns = INVENTORY_EXPORT_COLUMNS
        sheet_name = "Inventory"
    elif table_key_normalized == "par":
        rows = _filtered_par_rows(request.args, apply_filters=apply_filters)
        columns = PAR_EXPORT_COLUMNS
        sheet_name = "Par Locations"
    else:
        abort(404)

    hide_r_only = (request.args.get("hide_r_only") or "").strip().lower() == "true"
    if hide_r_only:
        rows = [row for row in rows if not _is_r_only_location(row)]

    column_mode = (request.args.get("column_mode") or "").strip().lower()
    requested_fields = _parse_column_selection(request.args.get("columns"))
    legacy_visible_param = request.args.get("visible_columns")
    if not requested_fields and legacy_visible_param:
        requested_fields = _parse_column_selection(legacy_visible_param)
        if not column_mode:
            column_mode = "visible"
    allowed_column_modes = {"all", "visible"} | CUSTOM_EXPORT_MODES
    if column_mode not in allowed_column_modes:
        column_mode = "all"

    if requested_fields:
        filtered_columns = _filter_export_columns(columns, requested_fields)
        if column_mode in CUSTOM_EXPORT_MODES:
            if not filtered_columns or len(filtered_columns) != len(requested_fields):
                abort(400, description="Requested columns are not available for export.")
            columns = filtered_columns
        elif filtered_columns:
            columns = filtered_columns
    elif column_mode in CUSTOM_EXPORT_MODES:
        abort(400, description="No columns selected for export.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    header_overrides = PRESET_HEADER_OVERRIDES.get(column_mode, {})
    worksheet.append([header_overrides.get(field, header) for header, field in columns])

    highlight_modes = {"inventory_setup", "par_setup"}
    should_highlight_notes = column_mode in highlight_modes
    notes_column_index: int | None = None
    if should_highlight_notes:
        for idx, (_, field_name) in enumerate(columns, start=1):
            if field_name == "notes":
                notes_column_index = idx
                break
        if notes_column_index is None:
            should_highlight_notes = False

    highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid") if should_highlight_notes else None

    for row_number, data_row in enumerate(rows, start=2):
        worksheet.append([_coerce_excel_value(data_row.get(field)) for _, field in columns])

        if should_highlight_notes and highlight_fill:
            notes_value = data_row.get("notes")
            has_notes = False
            if isinstance(notes_value, str):
                has_notes = notes_value.strip() != ""
            elif notes_value is not None:
                has_notes = str(notes_value).strip() != ""

            if has_notes:
                for col_idx in range(1, len(columns) + 1):
                    cell = worksheet.cell(row=row_number, column=col_idx)
                    cell.fill = highlight_fill

    worksheet.freeze_panes = "A2"
    if worksheet.max_row and worksheet.max_column:
        worksheet.auto_filter.ref = worksheet.dimensions

    max_row_for_width = min(worksheet.max_row, 200)
    for idx, column_cells in enumerate(worksheet.iter_cols(1, len(columns), 1, max_row_for_width), start=1):
        max_length = 0
        for cell in column_cells:
            value = cell.value
            length = len(str(value)) if value is not None else 0
            if length > max_length:
                max_length = length
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename_prefix = sheet_name.lower().replace(" ", "_")
    filename = f"{filename_prefix}_{timestamp}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/api/qty/<int:item_group>")
@login_required
def api_qty(item_group: int):
    """Return historical quantity time-series (AvailableQty) per (Item, Location)
    for a given item_group from the PLMQty view.

    Response structure:
    {
      "item_group": <int>,
      "series": [
          {"item": "12345", "location": "LOC1", "points": [
              {"t": "2025-08-01T00:00:00", "qty": 42}, ...
          ]}, ...
      ]
    }
    """
    # Query all rows for this item group ordered for stable client-side rendering
    stmt = (
        select(
            PLMQty.Item.label("item"),
            PLMQty.Location.label("location"),
            PLMQty.report_stamp.label("stamp"),
            PLMQty.AvailableQty.label("qty"),
            PLMQty.PLM_Zdate.label("z_date"),
        )
        .where(PLMQty.Item_Group == item_group)
        .order_by(PLMQty.Item, PLMQty.Location, PLMQty.report_stamp)
    )
    rows = db.session.execute(stmt).all()

    gl_map = {}
    if rows:
        gl_query = (
            select(PLMTrackerBase.Item, PLMTrackerBase.Location, PLMTrackerBase.Group_Locations)
            .where(PLMTrackerBase.Item_Group == item_group)
        )
        for item, location, group_loc in db.session.execute(gl_query).all():
            bucket = gl_map.setdefault(item, {})
            bucket[location] = group_loc
    series_map: dict[tuple[str, str], dict[str, object]] = {}
    for item, location, stamp, qty, z_date in rows:
        key = (item, location)
        bucket = series_map.setdefault(key, {"points": [], "z_date": None})
        points = bucket.setdefault("points", [])
        points.append({
            "t": stamp.isoformat() if stamp else None,
            "qty": int(qty) if qty is not None else None,
        })
        if z_date and bucket.get("z_date") is None:
            bucket["z_date"] = z_date.isoformat()
    series = [
        {
            "item": item_key,
            "location": loc_key,
            "group_location": gl_map.get(item_key, {}).get(loc_key) or gl_map.get(item_key, {}).get(None) or loc_key,
            "points": entry.get("points", []),
            "z_date": entry.get("z_date"),
        }
        for (item_key, loc_key), entry in series_map.items()
    ]

    return jsonify({
        "item_group": item_group,
        "series": series,
        "series_count": len(series),
        "point_count": sum(len(s["points"]) for s in series),
    })


@bp.route("/api/issue/<int:item_group>")
@login_required
def api_issue(item_group: int):
    """Return daily issue-out quantities per (Item, Location) for a given item_group.

    Response structure mirrors qty endpoint for easier client reuse.
    """
    stmt = (
        select(
            PLMDailyIssueOutQty.Item.label("item"),
            PLMDailyIssueOutQty.Location.label("location"),
            PLMDailyIssueOutQty.trx_date.label("stamp"),
            PLMDailyIssueOutQty.IssuedQty.label("qty"),
        )
        .where(PLMDailyIssueOutQty.Item_Group == item_group)
        .order_by(
            PLMDailyIssueOutQty.Item,
            PLMDailyIssueOutQty.Location,
            PLMDailyIssueOutQty.trx_date,
        )
    )
    rows = db.session.execute(stmt).all()

    gl_map = {}
    if rows:
        gl_query = (
            select(PLMTrackerBase.Item, PLMTrackerBase.Location, PLMTrackerBase.Group_Locations)
            .where(PLMTrackerBase.Item_Group == item_group)
        )
        for item, location, group_loc in db.session.execute(gl_query).all():
            bucket = gl_map.setdefault(item, {})
            bucket[location] = group_loc

    series_map = {}
    for item, location, stamp, qty in rows:
        key = (item, location)
        bucket = series_map.setdefault(key, [])
        bucket.append(
            {
                "t": stamp.isoformat() if stamp else None,
                "qty": int(qty) if qty is not None else None,
            }
        )

    series = [
        {
            "item": item_key,
            "location": loc_key,
            "group_location": gl_map.get(item_key, {}).get(loc_key)
            or gl_map.get(item_key, {}).get(None)
            or loc_key,
            "points": points,
        }
        for (item_key, loc_key), points in series_map.items()
    ]

    return jsonify(
        {
            "item_group": item_group,
            "series": series,
            "series_count": len(series),
            "point_count": sum(len(s["points"]) for s in series),
        }
    )
