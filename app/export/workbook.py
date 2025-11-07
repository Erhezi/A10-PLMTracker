from __future__ import annotations

from decimal import Decimal
from typing import Callable, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

Row = dict[str, object]
ColumnDef = Sequence[tuple[str, str]]


def _coerce_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def render_workbook(
    sheet_name: str,
    rows: Iterable[Row],
    columns: ColumnDef,
    *,
    header_overrides: dict[str, str] | None = None,
    highlight_notes: bool = False,
    highlight_row_predicate: Callable[[Row], bool] | None = None,
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]

    overrides = header_overrides or {}
    worksheet.append([overrides.get(field, header) for header, field in columns])

    highlight_fill = None
    notes_field_name = "notes"
    if highlight_notes:
        highlight_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for row_number, data_row in enumerate(rows, start=2):
        worksheet.append([_coerce_excel_value(data_row.get(field)) for _, field in columns])

        if highlight_fill:
            notes_value = data_row.get(notes_field_name)
            has_notes = False
            if isinstance(notes_value, str):
                has_notes = notes_value.strip() != ""
            elif notes_value is not None:
                has_notes = str(notes_value).strip() != ""

            should_highlight = True
            if highlight_row_predicate is not None:
                try:
                    should_highlight = bool(highlight_row_predicate(data_row))
                except Exception:
                    should_highlight = True

            if has_notes and should_highlight:
                for col_idx in range(1, len(columns) + 1):
                    cell = worksheet.cell(row=row_number, column=col_idx)
                    cell.fill = highlight_fill

    worksheet.freeze_panes = "A2"
    if worksheet.max_row and worksheet.max_column:
        worksheet.auto_filter.ref = worksheet.dimensions

    max_row_for_width = min(worksheet.max_row, 200)
    for idx, column_cells in enumerate(worksheet.iter_cols(1, len(columns), 1, max_row_for_width), start=1):
        max_length = 0
        for cell in column_cells:
            value = cell.value
            length = len(str(value)) if value is not None else 0
            if length > max_length:
                max_length = length
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

    return workbook


__all__ = ["render_workbook"]
